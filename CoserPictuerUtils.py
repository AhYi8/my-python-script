from openpyxl import Workbook, load_workbook
import requests, hashlib, uuid, random, string, time, os, re, openpyxl, shutil, redis
from PIL import Image, ImageDraw, ImageFont
from io import BytesIO
from datetime import datetime

class ImageUtils:
    smms_token: str = 'HkgOTOGlKPc5fkdQauKo5X5vpjj3hqVl'
    water: str = 'coser.21zys.com'
    pattern: str = 'coser.21zys.com-{Y}{m}{d}{h}{i}{s}'
    auto_increment: int = 1

    @staticmethod
    def find_first_image_with_text(directory, search_string):
        for root, _, files in os.walk(directory):
            for file in files:
                if search_string in file and file.lower().endswith(('.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.gif')):
                    return os.path.join(root, file)
        return None

    @staticmethod
    def add_watermark(image_path, watermark_text):
        """
        读取读取指定路径的图片，添加水印文本

        :param image_path: img图片路径
        :param watermark_text: 水印文本
        :return: img 流
        """
        image = Image.open(image_path)
        watermark_image = image.copy()
        draw = ImageDraw.Draw(watermark_image)

        font_size = int(min(image.size) * 0.1)  # 水印字体大小
        font = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", font_size)

        # 创建透明图层用于绘制水印
        watermark_layer = Image.new('RGBA', watermark_image.size, (0, 0, 0, 0))
        draw_watermark = ImageDraw.Draw(watermark_layer)

        # 计算文本的宽度和高度
        text_bbox = draw.textbbox((0, 0), watermark_text, font=font)
        text_width, text_height = text_bbox[2] - text_bbox[0], text_bbox[3] - text_bbox[1]

        # 计算水印的位置（居中）
        x = (watermark_layer.width - text_width) / 2
        y = (watermark_layer.height - text_height) / 2

        # 设置透明度（255为不透明）
        fill = (0, 0, 0, int(255 * 0.6))  # 黑色，透明度为0.4

        # 添加水印文本
        draw_watermark.text((x, y), watermark_text, font=font, fill=fill)

        # 旋转水印图层45度
        watermark_layer = watermark_layer.rotate(45, expand=True)

        # 创建一个与原始图像尺寸相同的空白图像
        full_watermark_layer = Image.new('RGBA', watermark_image.size, (0, 0, 0, 0))

        # 计算旋转后图层的位置以居中对齐
        paste_x = (full_watermark_layer.width - watermark_layer.width) // 2
        paste_y = (full_watermark_layer.height - watermark_layer.height) // 2

        # 将旋转后的水印图层粘贴到空白图像上
        full_watermark_layer.paste(watermark_layer, (paste_x, paste_y), watermark_layer)

        # 将水印图层与原始图像合并
        watermarked_image = Image.alpha_composite(watermark_image.convert('RGBA'), full_watermark_layer)

        watermarked_io = BytesIO()
        watermarked_image.convert('RGB').save(watermarked_io, format=image.format)
        watermarked_io.seek(0)
        return watermarked_io

    @staticmethod
    def add_watermark_to_image(image, watermark_text):
        watermark_image = image.copy()
        draw = ImageDraw.Draw(watermark_image)

        font_size = int(min(image.size) * 0.15)  # 水印字体大小
        font = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", font_size)

        # 创建透明图层用于绘制水印
        watermark_layer = Image.new('RGBA', watermark_image.size, (0, 0, 0, 0))
        draw_watermark = ImageDraw.Draw(watermark_layer)

        # 计算文本的宽度和高度
        text_bbox = draw.textbbox((0, 0), watermark_text, font=font)
        text_width, text_height = text_bbox[2] - text_bbox[0], text_bbox[3] - text_bbox[1]

        # 计算水印的位置（居中）
        x = (watermark_layer.width - text_width) / 2
        y = (watermark_layer.height - text_height) / 2

        # 设置透明度（255为不透明）
        fill = (0, 0, 0, int(255 * 0.6))  # 黑色，透明度为0.4

        # 添加水印文本
        draw_watermark.text((x, y), watermark_text, font=font, fill=fill)

        # 旋转水印图层45度
        watermark_layer = watermark_layer.rotate(45, expand=True)

        # 创建一个与原始图像尺寸相同的空白图像
        full_watermark_layer = Image.new('RGBA', watermark_image.size, (0, 0, 0, 0))

        # 计算旋转后图层的位置以居中对齐
        paste_x = (full_watermark_layer.width - watermark_layer.width) // 2
        paste_y = (full_watermark_layer.height - watermark_layer.height) // 2

        # 将旋转后的水印图层粘贴到空白图像上
        full_watermark_layer.paste(watermark_layer, (paste_x, paste_y), watermark_layer)

        # 将水印图层与原始图像合并
        watermarked_image = Image.alpha_composite(watermark_image.convert('RGBA'), full_watermark_layer)

        watermarked_io = BytesIO()
        watermarked_image.convert('RGB').save(watermarked_io, format=image.format)
        watermarked_io.seek(0)
        return watermarked_io

    @staticmethod
    def generate_filename(filename, pattern, auto_increment):
        """
        返回格式化的文件名，用于smms上传时的文件名格式化

        :param filename: 原文件名
        :param pattern: 格式化模板
        :param auto_increment: 自增序列
        :return: 格式化后的文件名
        """
        base, ext = filename.rsplit('.', 1)
        now = datetime.now()

        replacements = {
            '{Y}': now.strftime('%Y'),
            '{y}': now.strftime('%y'),
            '{m}': now.strftime('%m'),
            '{d}': now.strftime('%d'),
            '{h}': now.strftime('%H'),
            '{i}': now.strftime('%M'),
            '{s}': now.strftime('%S'),
            '{ms}': now.strftime('%f')[:3],
            '{timestamp}': str(int(now.timestamp())),
            '{md5}': hashlib.md5(str(uuid.uuid4()).encode()).hexdigest(),
            '{md5-16}': hashlib.md5(str(uuid.uuid4()).encode()).hexdigest()[:16],
            '{uuid}': str(uuid.uuid4()),
            '{str-6}': ''.join(random.choices(string.ascii_letters + string.digits, k=6)),
            '{filename}': base,
            '{auto}': str(auto_increment)
        }

        for key, value in replacements.items():
            pattern = pattern.replace(key, value)

        return pattern + '.' + ext

    @staticmethod
    def upload_to_smms_image(image_path):
        """
        上传本地文件到smms图床，期间需要格式化文件名，添加水印

        :param image_path: 本地图片路径
        :return: (原文件名, imgURL)
        """
        # 添加水印
        watermarked_image = ImageUtils.add_watermark(image_path, ImageUtils.water)

        # 获取原文件名去掉后缀部分
        original_filename = image_path.rsplit('/', 1)[-1].rsplit('.', 1)[0]

        new_filename = ImageUtils.generate_filename(image_path.rsplit('/', 1)[-1], ImageUtils.pattern,
                                                    ImageUtils.auto_increment)

        url = "https://sm.ms/api/v2/upload"
        headers = {'Authorization': ImageUtils.smms_token}
        files = {'smfile': (new_filename, watermarked_image)}

        response = requests.post(url, headers=headers, files=files)
        result = response.json()

        if response.status_code == 200 and result['success']:
            return original_filename, result['data']['url'], result['data']['delete']
        else:
            raise Exception(f"Upload to SMMS failed: {result.get('message', 'Unknown error')}")

    def upload_to_smms_by_image_url(image_url, original_filename):
        """
        上传网络文件到smms图床，期间需要格式化文件名，添加水印

        :param image_url: 本地图片路径
        :return: (原文件名, imgURL)
        """

        # 下载图片
        response = requests.get(image_url)
        if response.status_code != 200:
            raise Exception("Failed to download image.")

        image = Image.open(BytesIO(response.content))

        # 获取原始文件名后缀
        file_suffix = image_url.rsplit('/', 1)[-1].rsplit('.', 1)[1]

        # 添加水印
        watermarked_image_io = ImageUtils.add_watermark_to_image(image, ImageUtils.water)

        new_filename = ImageUtils.generate_filename(f'{original_filename}.{file_suffix}', ImageUtils.pattern,
                                                    ImageUtils.auto_increment)

        url = "https://sm.ms/api/v2/upload"
        headers = {'Authorization': ImageUtils.smms_token}
        files = {'smfile': (new_filename, watermarked_image_io)}

        response = requests.post(url, headers=headers, files=files)
        result = response.json()

        if response.status_code == 200 and result['success']:
            return original_filename, result['data']['url']
        else:
            raise Exception(f"Upload to SMMS failed: {result.get('message', 'Unknown error')}")

    @staticmethod
    def upload_to_smms(image_path, water, token, pattern, auto_increment):
        """
        上传本地文件到smms图床，期间需要格式化文件名，添加水印

        :param image_path: 本地图片路径
        :param water: 水印文本
        :param token: smms_token
        :param token: smms_token
        :param pattern: 文件名格式化模板
        :param auto_increment: 自增序列
        :return: (原文件名, imgURL)
        """
        # 添加水印
        watermarked_image = ImageUtils.add_watermark(image_path, water)

        # 获取原文件名去掉后缀部分
        original_filename = image_path.rsplit('/', 1)[-1].rsplit('.', 1)[0]

        new_filename = ImageUtils.generate_filename(image_path.rsplit('/', 1)[-1], pattern, auto_increment)

        url = "https://sm.ms/api/v2/upload"
        headers = {'Authorization': token}
        files = {'smfile': (new_filename, watermarked_image)}

        response = requests.post(url, headers=headers, files=files)
        result = response.json()

        if response.status_code == 200 and result['success']:
            return original_filename, result['data']['url']
        else:
            raise Exception(f"Upload to SMMS failed: {result.get('message', 'Unknown error')}")

    @staticmethod
    def upload_images_to_smms_from_directory(parent_path: str):
        success_results = {}
        failed_results = []
        allowed_formats = {'jpg', 'jpeg', 'png', 'gif', 'bmp'}

        # 遍历目录中的所有文件
        for root, dirs, files in os.walk(parent_path):
            for file in files:
                # 只处理图片文件
                if file.split('.')[-1].lower() in allowed_formats:
                    file_path = os.path.join(root, file)
                    original_filename: str = ''
                    # 调用 upload_to_smms 方法上传图片
                    try:
                        original_filename, imgURL, deleteUrl = ImageUtils.upload_to_smms_image(file_path)
                        success_results[original_filename] = imgURL  # 将结果存储到字典中
                    except Exception as e:
                        failed_results.append((original_filename, file_path))
                        print(f'图片上传失败：{original_filename}---->{file_path}---->{e}')
        return (success_results, failed_results)
        

def append_to_excel(image_list, excel_file):
    # 如果文件存在，则加载，否则创建一个新的 Excel 文件
    if os.path.exists(excel_file):
        workbook = load_workbook(excel_file)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        # 写入标题行
        sheet.append(['album', 'filename', 'absolute_path', 'image_url', 'delete_url'])

    # 将 image_list 中的数据逐行写入
    for image_info in image_list:
        sheet.append([
            image_info['album'], 
            image_info['filename'], 
            image_info['absolute_path'], 
            image_info['image_url'], 
            image_info['delete_url']
        ])

    # 保存 Excel 文件
    workbook.save(excel_file)

def get_images(directory, excel_file):
    image_list = []
    valid_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.bmp'}
    max_retries = 5  # 设置最大重试次数

    # 获取指定文件夹下的所有一级子文件夹
    for subdir in os.listdir(directory):
        subdir_path = os.path.join(directory, subdir)

        # 判断是否为一级子文件夹
        if os.path.isdir(subdir_path):
            album = subdir  # 保存当前一级子文件夹的名称到 album 变量

            # 递归遍历一级子文件夹及其所有子孙文件夹中的文件
            for root, dirs, files in os.walk(subdir_path):
                for file in files:
                    # 检查文件是否是图片（通过后缀名判断）
                    if os.path.splitext(file)[1].lower() in valid_extensions:
                        filename = file  # 获取图片文件名
                        absolute_path = os.path.abspath(os.path.join(root, file))  # 获取图片绝对路径
                        
                        image_info = {
                            'filename': filename,
                            'absolute_path': absolute_path,
                            'album': album,
                            'image_url': None,
                            'delete_url': None
                        }

                        # 上传图片到 SMMS，最多重试 5 次
                        retries = 0
                        while retries < max_retries:
                            try:
                                original_filename, image_url, delete_url = ImageUtils.upload_to_smms_image(absolute_path)
                                image_info['image_url'] = image_url
                                image_info['delete_url'] = delete_url
                                time.sleep(2)
                                break  # 上传成功，跳出循环
                            except Exception as e:
                                retries += 1
                                if retries >= max_retries:
                                    print(f"Failed to upload {filename} after {max_retries} retries. Error: {e}")
                                else:
                                    print(f"Retrying upload for {filename}... Attempt {retries}")
                                    time.sleep(2)  # 重试前等待 2 秒
                        print(image_info)
                        image_list.append(image_info)  # 将字典添加到 image_list 数组中
            
            # 遍历完当前一级子文件夹，将 image_list 追加写入 Excel
            append_to_excel(image_list, excel_file)

            # 清空 image_list，以便处理下一个 album
            image_list.clear()
            time.sleep(60)


def enable_proxy():
    os.environ['http_proxy'] = 'http://localhost:10809'
    os.environ['https_proxy'] = 'http://localhost:10809'
    print("全局代理已开启")

if __name__ == "__main__":
    enable_proxy()
    # 替换为你的目标文件夹路径
    target_directory = r"F:\BaiduNetdiskDownload\玉汇_thumbnail"
    excel_file = r"C:\Users\Administrator\Desktop\coser_album.xlsx"
    # 获取所有图片信息
    images = get_images(target_directory, excel_file)
