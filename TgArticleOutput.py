import requests, re, os, time, string, redis
from retrying import retry
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import List


class RedisUtils:
    # Redis 连接属性，用户提供
    host = "152.32.175.149"
    port = 6379
    db = 0
    password = 'Mh359687..'
    res_21zys_com_titles_key = 'res.21zys.com_titles'
    # 私有构造函数
    _redis_client = None

    @classmethod
    def _initialize_client(cls):
        if cls._redis_client is None:
            cls._redis_client = redis.StrictRedis(
                host=cls.host,
                port=cls.port,
                db=cls.db,
                password=cls.password,
                decode_responses=True
            )

    # ------------------ String 类型操作 ------------------
    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def set_string(cls, key: str, value: str):
        cls._initialize_client()
        return cls._redis_client.set(key, value)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def get_string(cls, key: str):
        cls._initialize_client()
        return cls._redis_client.get(key)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def del_key(cls, key: str):
        cls._initialize_client()
        return cls._redis_client.delete(key)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def update_string(cls, key: str, value: str):
        cls._initialize_client()
        return cls._redis_client.set(key, value)

    # ------------------ List 类型操作 ------------------
    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def push_list(cls, key: str, *values):
        cls._initialize_client()
        return cls._redis_client.rpush(key, *values)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def get_list(cls, key: str, start: int = 0, end: int = -1):
        cls._initialize_client()
        return cls._redis_client.lrange(key, start, end)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def pop_list(cls, key: str):
        cls._initialize_client()
        return cls._redis_client.lpop(key)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def list_length(cls, key: str):
        cls._initialize_client()
        return cls._redis_client.llen(key)

    # ------------------ Set 类型操作 ------------------
    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def add_set(cls, key: str, *values):
        cls._initialize_client()
        return cls._redis_client.sadd(key, *values)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def get_set(cls, key: str):
        cls._initialize_client()
        return cls._redis_client.smembers(key)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def rem_set(cls, key: str, *values):
        cls._initialize_client()
        return cls._redis_client.srem(key, *values)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def set_length(cls, key: str):
        cls._initialize_client()
        return cls._redis_client.scard(key)

    # ------------------ Hash 类型操作 ------------------
    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def set_hash(cls, key: str, field: str, value: str):
        cls._initialize_client()
        return cls._redis_client.hset(key, field, value)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def get_hash(cls, key: str, field: str):
        cls._initialize_client()
        return cls._redis_client.hget(key, field)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def get_all_hash(cls, key: str):
        cls._initialize_client()
        return cls._redis_client.hgetall(key)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def del_hash(cls, key: str, field: str):
        cls._initialize_client()
        return cls._redis_client.hdel(key, field)

    # ------------------ Zset 类型操作 ------------------
    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def add_zset(cls, key: str, score: float, value: str):
        cls._initialize_client()
        return cls._redis_client.zadd(key, {value: score})

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def get_zset(cls, key: str, start: int = 0, end: int = -1, withscores: bool = False):
        cls._initialize_client()
        return cls._redis_client.zrange(key, start, end, withscores=withscores)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def rem_zset(cls, key: str, *values):
        cls._initialize_client()
        return cls._redis_client.zrem(key, *values)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def zset_length(cls, key: str):
        cls._initialize_client()
        return cls._redis_client.zcard(key)


class ImageUtils:

    @staticmethod
    def download_image(image_url: str, title: str, save_path: str, retries: int = 5) -> bool:
        for attempt in range(retries):
            try:
                ext = os.path.splitext(image_url)[1]
                sanitized_title = title + ext
                image_path = os.path.join(save_path, sanitized_title)
                if not os.path.exists(image_path):
                    response = requests.get(image_url)
                    response.raise_for_status()
                    img = Image.open(BytesIO(response.content))
                    img.save(image_path)
                return True, sanitized_title
            except Exception as e:
                pass
        return False, ''


class ParseTgArticleUtils:
    @staticmethod
    def get_zh_vip_article(params):
        link_match = re.search(r'https://pan.quark.cn/s/[a-z0-9]{12}', params)
        link = link_match.group(0) if link_match else ''
        soup = BeautifulSoup(params, 'html.parser')
        for br in soup.find_all('br'):
            br.replace_with('\n')
        html_text = soup.get_text(separator='\n', strip=True)
        html_text_match = re.search(r"^(.*)\n(.*)[\s\S]*\n(标签：)?(((#\S+) *)*)$", html_text)
        if html_text_match:
            title = TgArticleUtils.sanitize_filename(TgArticleUtils.clean_title(html_text_match.group(1)))
            title_match = re.search(r'^[\[【](.*)[】\]]$', title)
            title = title_match.group(1) if title_match else title
            title = re.sub(r'\s*(\d+)\s*', r'\1', title)
            description = html_text_match.group(2)
            tags = html_text_match.group(4).split('#')
            tags = [tag.strip() for tag in tags if tag.strip() not in TgArticleUtils.tag_remove_keys]
            tag = ','.join(tags).strip(',').replace(',,', ',')
            return title, description, link, "N", tag
        return None

    @staticmethod
    def get_other_tg_quark_article(params):
        title_regex = r'(名称|资源标题)：(.+)'
        description_regex = r'(描述|资源描述)：([\s\S]*)(链接：)'
        size_regex = r'大小：(.+)'
        tag_regex = r'标签：(.+)'

        link_match = re.search(r'https://pan.quark.cn/s/[a-z0-9]{12}', params)
        link = link_match.group(0) if link_match else ''
        soup = BeautifulSoup(params, 'html.parser')
        html_text = soup.get_text(separator='\n', strip=True)
        title_match = re.search(title_regex, html_text)
        description_match = re.search(description_regex, html_text)
        size_match = re.search(size_regex, html_text)
        tag_match = re.search(tag_regex, html_text)

        title = TgArticleUtils.sanitize_filename(
            TgArticleUtils.clean_title(title_match.group(2))) if title_match else ''
        title_match = re.search(r'^[\[【](.*)[】\]]$', title)
        title = title_match.group(1) if title_match else title
        title = re.sub(r'\s*(\d+)\s*', r'\1', title)
        description = description_match.group(2).strip() if description_match else ''
        size = size_match.group(1) if size_match else ''
        tags = tag_match.group(1).replace(' ', '').strip('#').split('#') if tag_match else []
        tags = [tag.strip() for tag in tags if tag.strip() not in TgArticleUtils.tag_remove_keys]
        tag = ','.join(tags).strip(',').replace(',,', ',')
        return title, description, link, size, tag


class FileUtils:

    @staticmethod
    def read_file(path: str, is_strip: bool = False, mode: str = 'r', encoding: str = 'u8'):
        if path is not None and len(path.strip()) != 0 and os.path.exists(path):
            try:
                with open(path, mode=mode, encoding=encoding) as rf:
                    data_list = rf.readlines()
                if is_strip:
                    data_list = [data.strip() for data in data_list]
                return data_list
            except BaseException as e:
                print(f"FileUtils.read_file(): {e}")
                return None
        else:
            return None

    @staticmethod
    def write_file(path: str, data_list, mode: str = 'w', encoding: str = 'u8'):
        if path is not None and len(path.strip()) != 0:
            parent_path = os.path.split(path)[0]
            if not os.path.exists(parent_path):
                os.makedirs(parent_path)
            try:
                write_list = []
                for data in data_list:
                    if not data.endswith("\n"):
                        data = data + '\n'
                    write_list.append(data)

                with open(path, mode=mode, encoding=encoding) as wf:
                    wf.writelines(write_list)
            except BaseException as e:
                print(f"FileUtils.read_file(): {e}")

    @staticmethod
    def remove_duplicate_logs(log_path: str = os.path.join(os.getcwd(), 'file', 'logs.txt')):
        logs: set = {log for log in FileUtils.read_file(log_path, is_strip=True) if '跳过采集' not in log and '文章不存在' not in log}
        FileUtils.write_file(log_path, logs)


class TgArticleUtils:
    get_tg_article_map = {
        'VIP资源共享': ParseTgArticleUtils.get_zh_vip_article,
        'other': ParseTgArticleUtils.get_other_tg_quark_article
    }
    invalid_chars = r'\/:*?"<>|'
    ads_articles = {'TG必备的搜索引擎，极搜帮你精准找到，想要的群组、频道、音乐 、视频'}
    tag_remove_keys = ['中国', '课程', '中国', '教程', '夸克', '夸克网盘', 'quark', '资源', '知识', '学习']
    exists_titles = RedisUtils.get_set(RedisUtils.res_21zys_com_titles_key)

    @staticmethod
    def sanitize_filename(title: str) -> str:
        return ''.join([c if c not in TgArticleUtils.invalid_chars else '_' for c in title])

    @staticmethod
    def clean_title(title: str) -> str:
        cleaned_title = title.replace('&#xFEFF;', '').replace('\ufeff', '')
        return cleaned_title.strip()

    @staticmethod
    def fetch_page(url: str, retries: int = 5, delay: int = 2) -> str:
        proxies = {
            "http": "http://localhost:10809",
            "https": "http://localhost:10809",
        }
        """获取页面内容，并在失败时进行重试，最多重试 `retries` 次，每次重试间隔 `delay` 秒"""
        for attempt in range(1, retries + 1):
            try:
                response = requests.get(url, proxies=proxies)
                response.raise_for_status()  # 如果响应状态码不是200，抛出异常
                return response.content
            except requests.RequestException as e:
                if attempt < retries:
                    time.sleep(delay)  # 等待几秒钟再重试
                else:
                    return None

    @staticmethod
    def get_tg_article_content(html_content: str):
        """
        获取 tg 的 author，首图链接，HTML 内容
        :param html_content:
        :return: author, image_url, html_content
        """
        author = image_url = content = ''
        try:
            soup = BeautifulSoup(html_content, 'html.parser')
            error = soup.find('div', class_="tgme_widget_message_error")
            if error:
                return '', '', 'Post not found'
            # 3. 获取作者信息
            author = soup.find('div', class_='tgme_widget_message_author').get_text(strip=True)

            # 4. 获取图片链接
            image_url = soup.find('a', class_='tgme_widget_message_photo_wrap')['style']
            image_url = image_url.split('url(')[-1].split(')')[0].strip('"').strip("'")
            
            # 5. 获取 HTML 内容
            content = soup.find('div', class_='tgme_widget_message_text').decode_contents()
        except Exception as e:
            pass
        
        return author, image_url, content

    @staticmethod
    def get_tg_article(author: str, html_content: str):
        """
        获取 tg_article 文章的通用接口，调用实现类方法既可以
        :param author: 作者
        :param html_content: html
        :return: (title, content, link, size, tag)
        """
        author = author if author in TgArticleUtils.get_tg_article_map.keys() else 'other'
        return TgArticleUtils.get_tg_article_map[author](html_content)

    
    @staticmethod
    def process_url(base_url: str, image_save_path: str, excel_file: str, retries: int = 3, delay: int = 2):
        cwd = os.getcwd()
        url = f"{base_url}?embed=1&mode=tme"
        author = title = renamed_image = content = link = size = tag = ''
        success_str = '失败'
        logs_path = os.path.join(cwd, 'file', 'logs.txt')
            
        for attempt in range(1, retries + 1):
            try:
                # 1. Fetch the page content
                html_content = TgArticleUtils.fetch_page(url)
                if html_content is None:
                    FileUtils.write_file(logs_path, [f"TgArticleUtils_无法获取页面内容: {url}"], 'a')
                    raise Exception(f"无法获取页面内容: {url}")

                # 2. Parse meta tags (image and description)
                author, image_url, html_content = TgArticleUtils.get_tg_article_content(html_content)
                if html_content == 'Post not found':
                    FileUtils.write_file(logs_path, [f"TgArticleUtils_文章不存在: {url}"], 'a')
                    return None
                if not image_url or not html_content:
                    FileUtils.write_file(logs_path, [f"TgArticleUtils_无法解析页面内容: {url}"], 'a')
                    raise Exception(f"无法解析页面内容: {url}")
                    
                for ads in TgArticleUtils.ads_articles:
                    if ads in html_content:
                        FileUtils.write_file(logs_path, [f"TgArticleUtils_广告文章，跳过采集: {url}"], 'a')
                        return None

                # 3. Apply regex to extract information
                result = TgArticleUtils.get_tg_article(author, html_content)
                if not result:
                    FileUtils.write_file(logs_path, [f"TgArticleUtils_正则解析失败: {url}"], 'a')
                    return base_url, author, success_str, title, renamed_image, content, link, size, tag
                    
                title, content, link, size, tag = result
                if not title or not link:
                    FileUtils.write_file(logs_path, [f"TgArticleUtils_正则解析失败: {url}"], 'a')
                    return base_url, author, success_str, title, renamed_image, content, link, size, tag

                # 4. Download the image
                success, renamed_image = ImageUtils.download_image(image_url, title, image_save_path)
                if not success:
                    FileUtils.write_file(logs_path, [f"TgArticleUtils_图片下载失败: {url}"], 'a')
                    raise Exception(f"图片下载失败: {url}")

                success_str = "成功" if success else "失败"
                if TgArticleUtils.exists_titles and title in TgArticleUtils.exists_titles:
                    FileUtils.write_file(logs_path, [f"TgArticleUtils_文章已发布：{title}，跳过采集：{url}"], 'a')
                    return None
                # 5. Return all data for appending to Excel
                # URL, Author, Success, Title, Image, Description, Status, Tags, Categories, Price, Link, Size
                return base_url, author, success_str, title, None, content, '', tag, None, 0, link, size
            except Exception as e:
                print(f"处理 URL {base_url} 时发生错误: {e}")
                if attempt == retries:
                    return None
                else:
                    time.sleep(delay)
                

    @staticmethod
    def append_to_excel(file_name: str, data: tuple):
        if not os.path.exists(file_name):
            wb = Workbook()
            ws = wb.active
            ws.title = "TgArticle"
            ws.append(['URL', 'Author', 'Success', 'Title', 'Image', 'Description', 'Status', 'Tags', 'Categories', 'Price', 'Link', 'Size'])
            wb.save(file_name)
            
        wb = load_workbook(file_name)
        ws = wb.active
        ws.append(data)
        wb.save(file_name)
        
    @staticmethod
    def tg_article_output(concurrency: int = None):
        cwd = os.getcwd()
        image_save_path = os.path.join(cwd, 'image')
        if not os.path.exists(image_save_path):
            os.makedirs(image_save_path)
        excel_file = os.path.join(cwd, 'file', 'tg_articles.xlsx')
        if os.path.exists(excel_file):
            os.remove(excel_file)
        urls = FileUtils.read_file(os.path.join(cwd, 'file', 'un_publish_articles.txt'), is_strip=True)

        if not urls:
            print("未找到有效的 URL，检查 urls.txt 文件")
            return

        # 自动检测 CPU 核心数，若用户未指定并发度则自动设置为 CPU 核心数的 2 倍
        if concurrency is None:
            cpu_count = os.cpu_count()
            concurrency = cpu_count * 2 if cpu_count else 4  # 保障至少有 4 个并发线程

        print(f"使用 {concurrency} 个并发线程进行处理")

        with ThreadPoolExecutor(max_workers=concurrency) as executor:
            future_to_url = {executor.submit(TgArticleUtils.process_url, url, image_save_path, excel_file): url for url in
                             urls}
            total = len(urls)
            zfill_size = len(str(total))
            index = 1

            for future in as_completed(future_to_url):
                result = future.result()
                if result:
                    TgArticleUtils.append_to_excel(excel_file, result)
                    url = future_to_url[future]
                    print(f"{str(index).zfill(zfill_size)}/{total}-->处理完成: {url}")
                index += 1
        FileUtils.remove_duplicate_logs()
        input('\n\n回车结束程序（enter）')
    






def enable_proxy():
    os.environ['http_proxy'] = 'http://localhost:10809'
    os.environ['https_proxy'] = 'http://localhost:10809'
    print("全局代理已开启")


if __name__ == "__main__":
    # 开启全局代理
    enable_proxy()
    # 如果不传递并发度，会自动检测CPU并设置并发数
    TgArticleUtils.tg_article_output()
    #FileUtils.remove_duplicate_logs()