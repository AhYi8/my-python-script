import os, logging, requests, redis, pytz, re, random, collections.abc, time, schedule
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from dateutil import parser
from retrying import retry
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from wordpress_xmlrpc import Client, WordPressPost, ServerConnectionError
from wordpress_xmlrpc.methods.posts import GetPosts, NewPost
from wordpress_xmlrpc.methods import taxonomies
from openpyxl import Workbook
from typing import List
from urllib.parse import urlparse

collections.Iterable = collections.abc.Iterable

# 创建日志目录
log_dir = os.path.join(os.getcwd(), 'file')
if not os.path.exists(log_dir):
    os.makedirs(log_dir)

# 创建一个 FileHandler，并设置编码为 'utf-8'
log_file = os.path.join(log_dir, 'log.txt')
file_handler = logging.FileHandler(log_file, encoding='utf-8')

# 创建一个日志格式器
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

# 将格式器应用于文件处理器
file_handler.setFormatter(formatter)

# 创建一个 StreamHandler 用于输出到控制台
console_handler = logging.StreamHandler()
console_handler.setFormatter(formatter)

# 获取根日志记录器并配置
logger = logging.getLogger()
logger.setLevel(logging.INFO)  # 设置日志级别
logger.addHandler(file_handler)  # 添加文件处理器
logger.addHandler(console_handler)  # 添加控制台处理器


class Article:
    title: str
    content: str
    date: datetime
    post_status: str
    terms_names: {}
    custom_fields: []

    def __init__(self):
        super()


class WordpressUtils:
    wp = Client(
        r'http://res.21zys.com/xmlrpc_Mh359687...php',
        '21zys',
        'Mh359687..'
    )

    @staticmethod
    def get_all_tag_name():
        return {tag.name for tag in WordpressUtils.wp.call(taxonomies.GetTerms('post_tag'))}

    @staticmethod
    def get_all_post(number=100, offset=0, limit=0):
        all_posts = []
        while True:
            try:
                # 获取一部分文章，使用 offset 和 number 控制
                posts = WordpressUtils.wp.call(GetPosts({'number': number, 'offset': offset}))
                if not posts:
                    break  # 如果没有返回文章，说明已经获取完毕
                all_posts.extend(posts)
                if limit != 0 and len(all_posts) >= limit:
                    break
                offset += number  # 更新 offset，以便获取下一部分文章
            except ServerConnectionError as e:
                logging.error(f"Connection error: {e}")
                break
            except Exception as e:
                logging.error(f"An error occurred: {e}")
                break
        return all_posts

    @staticmethod
    def post_article(article: Article):
        try:
            post = WordPressPost()
            post.title = article.title
            if article.date:
                post.date = article.date
                post.date_modified = article.date
            post.content = article.content
            post.post_status = article.post_status
            post.terms_names = article.terms_names
            post.custom_fields = article.custom_fields
            post.comment_status = 'open'
            post.id = WordpressUtils.wp.call(NewPost(post))
            return post.id
        except Exception as e:
            logging.error(f"发布文章失败：{e}")
            return None

    @staticmethod
    def post_articles(articles: List[Article]):
        total = len(articles)
        zfill_size = len(str(total))
        index = 1
        for article in articles:
            logging.info(f"{str(index).zfill(zfill_size)}/{total}-->正在发布：{article.title}")
            print(f"{str(index).zfill(zfill_size)}/{total}-->正在发布：{article.title}")
            index += 1
            WordpressUtils.post_article(article)


class DateUtils:
    @staticmethod
    def get_current_date():
        """获取当前日期"""
        return datetime.now().date()

    @staticmethod
    def get_current_datetime():
        """获取当前日期时间"""
        return datetime.now()

    @staticmethod
    def get_current_datetime_str(fmt: str = "%Y-%m-%d %H:%M:%S") -> str:
        """返回当前日期时间字符串，格式: 2024-09-08 01:32:08"""
        return datetime.now().strftime(fmt)

    @staticmethod
    def str_to_date(date_str: str, fmt: str = "%Y-%m-%d") -> datetime.date:
        """将字符串转换为日期"""
        return datetime.strptime(date_str, fmt).date()

    @staticmethod
    def date_to_str(date: datetime.date, fmt: str = "%Y-%m-%d") -> str:
        """将日期转换为字符串"""
        return date.strftime(fmt)

    @staticmethod
    def datetime_to_str(date: datetime, fmt: str = "%Y-%m-%d %H:%M:%S"):
        return date.strftime(fmt)

    @staticmethod
    def iso_str_to_datetime(iso_str: str) -> datetime:
        """将带时区的 ISO 8601 格式时间字符串转换为 datetime 对象"""
        return parser.parse(iso_str)

    @staticmethod
    def add_days(date: datetime.date, days: int) -> datetime.date:
        """在指定日期基础上增加或减少天数"""
        return date + timedelta(days=days)

    @staticmethod
    def add_months(date: datetime.date, months: int) -> datetime.date:
        """在指定日期基础上增加或减少月份"""
        return date + relativedelta(months=months)

    @staticmethod
    def days_between(date1: datetime.date, date2: datetime.date) -> int:
        """计算两个日期之间的天数差"""
        return (date2 - date1).days

    @staticmethod
    def date_to_datetime(date_obj: datetime.date) -> datetime:
        """将 date 对象转换为 datetime 对象"""
        return datetime.combine(date_obj, datetime.min.time())

    @staticmethod
    def datetime_to_date(datetime_obj: datetime) -> datetime.date:
        """将 datetime 对象转换为 date 对象"""
        return datetime_obj.date()

    @staticmethod
    def is_valid_date(date_str: str, fmt: str = "%Y-%m-%d") -> bool:
        """判断字符串是否为合法的日期格式"""
        try:
            datetime.strptime(date_str, fmt)
            return True
        except ValueError:
            return False

    @staticmethod
    def get_current_year():
        """获取当前年份"""
        return datetime.now().year

    @staticmethod
    def get_current_month():
        """获取当前月份"""
        return datetime.now().month

    @staticmethod
    def get_current_day():
        """获取当前日期中的日"""
        return datetime.now().day

    @staticmethod
    def get_timezone_aware_datetime(timezone_str: str = 'UTC'):
        """返回指定时区的当前时间"""
        tz = pytz.timezone(timezone_str)
        return datetime.now(tz)

    @staticmethod
    def is_before(date1: datetime, date2: datetime) -> bool:
        """判断 date1 是否在 date2 之前"""
        return date1 < date2

    @staticmethod
    def is_after(date1: datetime, date2: datetime) -> bool:
        """判断 date1 是否在 date2 之后"""
        return date1 > date2

    @staticmethod
    def is_same(date1: datetime, date2: datetime) -> bool:
        """判断 date1 和 date2 是否相同"""
        return date1 == date2

    @staticmethod
    def to_naive(datetime_obj: datetime) -> datetime:
        """将带时区的 datetime 转换为 naive datetime（移除时区信息）"""
        if datetime_obj.tzinfo is not None:
            return datetime_obj.replace(tzinfo=None)
        return datetime_obj


class RedisUtils:
    # Redis 连接属性，用户提供
    host = "152.32.175.149"
    port = 6379
    db = 0
    password = 'Mh359687..'
    vip_91_article_links = 'vip_91_article_links'
    # 私有构造函数
    _redis_client = None

    @classmethod
    def _initialize_client(cls):
        if cls._redis_client is None:
            cls._redis_client = redis.StrictRedis(
                host=cls.host,
                port=cls.port,
                db=cls.db,
                password=cls.password,
                decode_responses=True
            )

    # ------------------ String 类型操作 ------------------
    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def set_string(cls, key: str, value: str):
        cls._initialize_client()
        return cls._redis_client.set(key, value)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def get_string(cls, key: str):
        cls._initialize_client()
        return cls._redis_client.get(key)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def del_key(cls, key: str):
        cls._initialize_client()
        return cls._redis_client.delete(key)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def update_string(cls, key: str, value: str):
        cls._initialize_client()
        return cls._redis_client.set(key, value)

    # ------------------ List 类型操作 ------------------
    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def push_list(cls, key: str, *values):
        cls._initialize_client()
        return cls._redis_client.rpush(key, *values)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def get_list(cls, key: str, start: int = 0, end: int = -1):
        cls._initialize_client()
        return cls._redis_client.lrange(key, start, end)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def pop_list(cls, key: str):
        cls._initialize_client()
        return cls._redis_client.lpop(key)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def list_length(cls, key: str):
        cls._initialize_client()
        return cls._redis_client.llen(key)

    # ------------------ Set 类型操作 ------------------
    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def add_set(cls, key: str, *values):
        cls._initialize_client()
        return cls._redis_client.sadd(key, *values)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def get_set(cls, key: str):
        cls._initialize_client()
        return cls._redis_client.smembers(key)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def rem_set(cls, key: str, *values):
        cls._initialize_client()
        return cls._redis_client.srem(key, *values)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def set_length(cls, key: str):
        cls._initialize_client()
        return cls._redis_client.scard(key)

    # ------------------ Hash 类型操作 ------------------
    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def set_hash(cls, key: str, field: str, value: str):
        cls._initialize_client()
        return cls._redis_client.hset(key, field, value)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def get_hash(cls, key: str, field: str):
        cls._initialize_client()
        return cls._redis_client.hget(key, field)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def get_all_hash(cls, key: str):
        cls._initialize_client()
        return cls._redis_client.hgetall(key)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def del_hash(cls, key: str, field: str):
        cls._initialize_client()
        return cls._redis_client.hdel(key, field)

    # ------------------ Zset 类型操作 ------------------
    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def add_zset(cls, key: str, score: float, value: str):
        cls._initialize_client()
        return cls._redis_client.zadd(key, {value: score})

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def get_zset(cls, key: str, start: int = 0, end: int = -1, withscores: bool = False):
        cls._initialize_client()
        return cls._redis_client.zrange(key, start, end, withscores=withscores)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def rem_zset(cls, key: str, *values):
        cls._initialize_client()
        return cls._redis_client.zrem(key, *values)

    @classmethod
    @retry(stop_max_attempt_number=5, wait_fixed=2000)
    def zset_length(cls, key: str):
        cls._initialize_client()
        return cls._redis_client.zcard(key)


class FileUtils:
    @staticmethod
    def read_file(path: str, is_strip: bool = False, mode: str = 'r', encoding: str = 'u8'):
        if path is not None and len(path.strip()) != 0 and os.path.exists(path):
            try:
                with open(path, mode=mode, encoding=encoding) as rf:
                    data_list = rf.readlines()
                if is_strip:
                    data_list = [data.strip() for data in data_list]
                return data_list
            except BaseException as e:
                print(f"FileUtils.read_file(): {e}")
                return None
        else:
            return None

# 配置日志
class Vip91ChuangYeUtils:
    tag_names = WordpressUtils.get_all_tag_name()

    @staticmethod
    def convert_cookie_to_dict(cookie_str):
        """
        将字符串类型的cookie转换为requests模块需要的字典格式
        :param cookie_str: 字符串类型的cookie (格式为 "key1=value1; key2=value2; ...")
        :return: 字典类型的cookie
        """
        cookies = {}
        # 通过分号分割各个键值对
        cookie_items = cookie_str.split(';')

        # 遍历每个键值对，去掉空白并将它们添加到字典中
        for item in cookie_items:
            key, value = item.strip().split('=', 1)
            cookies[key] = value

        return cookies

    @staticmethod
    def fetch_url(url: str, headers: dict = None, retries=1):
        """发送HTTP请求，默认重试1次"""
        # logging.info(f"Fetching URL: {url}")
        attempt = 0
        while attempt <= retries:
            try:
                if headers:
                    response = requests.get(url, headers=headers)
                else:
                    response = requests.get(url)
                response.raise_for_status()
                return response.text
            except requests.RequestException as e:
                logging.error(f"Error fetching URL: {url}, attempt {attempt + 1}, Error: {e}")
                if attempt == retries:
                    return None  # 超过最大重试次数，返回None
                attempt += 1

    @staticmethod
    def publish_vip_91_chuangye_article(base_url,
                                        cookies,
                                        is_append_xlsx: bool = True,
                                        xlsx_file_path: str = os.path.join(os.getcwd(), "file", "vip_91chuangye_com.xlsx")):
        if base_url.endswith("/"):
            base_url = base_url[:-1]

        page = 1
        published_article_links = RedisUtils.get_set(RedisUtils.vip_91_article_links)
        flag = True
        stop_date = DateUtils.date_to_datetime(DateUtils.add_days(DateUtils.get_current_date(), -3))

        while flag:
            article_metas = []
            results = []
            if page != 1:
                url = f"{base_url}/page/{page}"
            else:
                url = base_url
            html_content = Vip91ChuangYeUtils.fetch_url(url)

            if not html_content:
                break  # 如果请求失败，退出循环

            soup = BeautifulSoup(html_content, 'html.parser')

            post_wrapper = soup.find('div', class_='row posts-wrapper scroll')

            item_list = post_wrapper.find_all('div', recursive=False) if post_wrapper else []

            for item in item_list:
                try:

                    # 提取title和link
                    entry_wrapper = item.find('div', class_='entry-wrapper')
                    title_tag = entry_wrapper.find('h2', class_='entry-title').find('a')
                    title = title_tag.text.strip()
                    link = title_tag['href']

                    if published_article_links and link in published_article_links:
                        continue

                    source_url, source_pwd, article_content = Vip91ChuangYeUtils.get_vip_practical_project_content(link, title, cookies)

                    # 提取cover
                    cover_tag = item.find('div', class_='entry-media').find('img')
                    cover = cover_tag['data-src'] if cover_tag else None

                    # 提取category
                    category_tags = entry_wrapper.find('span', class_='meta-category-dot').find_all('a')
                    category = ",".join(a.text.strip() for a in category_tags)

                    publish_date_tags = entry_wrapper.find('span', class_='meta-date').find('time')
                    publish_date = DateUtils.to_naive(DateUtils.iso_str_to_datetime(
                        publish_date_tags['datetime'])) if publish_date_tags else DateUtils.get_current_datetime()

                    # if DateUtils.is_before(publish_date, stop_date):
                    #     flag = False
                    #     break

                    # logging.info(f"Link: {link}, Title: {title}, Cover: {cover}, Category: {category}, Publish_date: {publish_date}")
                    results.append([link, title, cover, category, article_content, source_url, source_pwd,DateUtils.datetime_to_str(publish_date)])
                    article_metas.append([link, title, cover, category, article_content, source_url, source_pwd, publish_date])
                except Exception as e:
                    logging.error(f"Error processing article: {e}")
                    continue
            # 下一页
            page += 1
            if is_append_xlsx:
                # 将结果追加到xlsx文件中
                Vip91ChuangYeUtils.append_to_xlsx(xlsx_file_path, results)

            if article_metas:
                for article_meta in article_metas:
                    link, title, cover, category, content, source_url, source_pwd, publish_date = article_meta
                    article: Article = Vip91ChuangYeUtils.article_meta_article(title, content, category, source_url, source_pwd, publish_date)
                    logging.info(f"正在采集发布：{link}-->{title}")
                    post_id = WordpressUtils.post_article(article)
                    if post_id:
                        RedisUtils.add_set(RedisUtils.vip_91_article_links, link)

    @staticmethod
    def get_vip_practical_project_content(url: str, title: str, cookies: str):
        # 自定义请求头
        if 'vip.91chuangye' in url:
            headers = {
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
                'Accept-Encoding': 'gzip, deflate, br, zstd',
                'Accept-Language': 'zh-CN,zh;q=0.9',
                'Cache-Control': 'max-age=0',
                'Connection': 'keep-alive',
                'Cookie': cookies,
                'Host': 'vip.91chuangye.cn',
                'Referer': 'https://vip.91chuangye.cn/',
                'Sec-Fetch-Dest': 'document',
                'Sec-Fetch-Mode': 'navigate',
                'Sec-Fetch-Site': 'same-origin',
                'Sec-Fetch-User': '?1',
                'Upgrade-Insecure-Requests': '1',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36',
                'sec-ch-ua': '"Google Chrome";v="129", "Not=A?Brand";v="8", "Chromium";v="129"',
                'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"Windows"'
            }
        elif 'bbs.abab9' in url:
            # 使用urlparse解析URL
            parsed_url = urlparse(url)
            path = parsed_url.path  # 获取URL的路径

            # 定义请求头
            headers = {
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
                'Accept-Encoding': 'gzip, deflate, br, zstd',
                'Accept-Language': 'zh-CN,zh;q=0.9',
                'Cache-Control': 'max-age=0',
                'Cookie': cookies,
                'Priority': 'u=0, i',
                'Sec-Ch-Ua': '"Google Chrome";v="129", "Not=A?Brand";v="8", "Chromium";v="129"',
                'Sec-Ch-Ua-Mobile': '?0',
                'Sec-Ch-Ua-Platform': '"Windows"',
                'Sec-Fetch-Dest': 'document',
                'Sec-Fetch-Mode': 'navigate',
                'Sec-Fetch-Site': 'none',
                'Sec-Fetch-User': '?1',
                'Upgrade-Insecure-Requests': '1',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36'
            }
        else:
            headers = {
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
                'Accept-Encoding': 'gzip, deflate, br, zstd',
                'Accept-Language': 'zh-CN,zh;q=0.9',
                'Cache-Control': 'max-age=0',
                'Cookie': cookies,  # 假设 `cookies` 是一样的
                'Sec-Fetch-Dest': 'document',
                'Sec-Fetch-Mode': 'navigate',
                'Sec-Fetch-User': '?1',
                'Upgrade-Insecure-Requests': '1',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36',
                'Sec-Ch-Ua': '"Google Chrome";v="129", "Not=A?Brand";v="8", "Chromium";v="129"',
                'Sec-Ch-Ua-Mobile': '?0',
                'Sec-Ch-Ua-Platform': '"Windows"'
            }


        html_content = Vip91ChuangYeUtils.fetch_url(url, headers)
        soup = BeautifulSoup(html_content, 'html.parser')
        try:
            content = soup.find('div', class_='entry-content u-text-format u-clearfix').decode_contents()
            return Vip91ChuangYeUtils.handle_content(title, content)
        except Exception as e:
            logging.error(f"Error processing article: {e}")
        return None

    @staticmethod
    def handle_content(title, content: str):
        # 匹配百度网盘链接
        baidu_pattern1 = r'(?:https?://)?\b(e?yun|pan)\.baidu\.com/[sj]/([\w\-]{5,})(?!\.)'
        baidu_pattern2 = r'(?:https?://)?\b(e?yun|pan)\.baidu\.com/(?:share|wap)/init\?surl=([\w\-]{5,})(?!\.)'
        baidu_pattern3 = r'http://pan\.baidu\.com/share/link\?shareid=[0-9]*&amp;uk=[0-9]*'
        url = None
        article_content = None
        try:
            article_content = re.search(r"""([\s\S]*)<div class="ripay-content""", content).group(1).replace(
                """ decoding=\"async\"""", "")
        except Exception as e:
            logging.error(f"{title}：{e}")
        # 检查百度网盘链接
        match = re.search(f"{baidu_pattern1}|{baidu_pattern2}|{baidu_pattern3}", content, re.IGNORECASE)
        if match:
            # 百度网盘链接检测
            url = match.group(0)

        pwd_match = re.search(r"\?pwd=(.{4})|提取码[:：](.{4})", content)
        pwd = pwd_match.group(1) if pwd_match else ""
        return url, pwd, article_content

    @staticmethod
    def article_meta_article(title, content, category, source_url, source_pwd, publish_date):
        article = Article()
        article.title = title
        article.content = content.replace(r"<img", r"<img class='aligncenter'")
        article.date = publish_date
        article.post_status = "publish"
        tags = {tag_name for tag_name in Vip91ChuangYeUtils.tag_names if
                tag_name in article.content or tag_name in title}
        terms_names = {'post_tag': list(tags), 'category': category.split(",")}
        article.terms_names = terms_names

        cao_downurl_new = [{'name': title, 'url': source_url, 'pwd': source_pwd}]
        keywords = []
        keywords.extend(list(tags))
        keywords.extend(category)
        custom_fields = [
            {'key': 'cao_price', 'value': "99.9"},
            {'key': 'cao_vip_rate', 'value': "0.6"},
            {'key': 'cao_is_boosvip', 'value': "1"},
            {'key': 'cao_close_novip_pay', 'value': 0},
            {'key': 'cao_paynum', 'value': "0"},
            {'key': 'cao_status', 'value': "1"},
            {'key': 'cao_downurl_new', 'value': cao_downurl_new},
            {'key': 'cao_info', 'value': ""},
            {'key': 'cao_demourl', 'value': ""},
            {'key': 'cao_diy_btn', 'value': ""},
            {'key': 'cao_video', 'value': ""},
            {'key': 'cao_is_video_free', 'value': ""},
            {'key': 'video_url_new', 'value': ""},
            {'key': 'post_titie', 'value': ""},
            {'key': 'keywords', 'value': ','.join(keywords)},
            {'key': 'description', 'value': BeautifulSoup(content, "html.parser").get_text(strip=True)},
            {'key': 'views', 'value': str(random.randint(300, 500))}
        ]
        article.custom_fields = custom_fields
        return article

    @staticmethod
    def append_to_xlsx(file_path, data):
        """将提取到的结果追加到现有的xlsx文件中，如果不存在则创建新文件"""
        if not os.path.exists(file_path):
            # 如果文件不存在，创建新文件并添加表头
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            ws.append(["Link", "Title", "Cover", "Category", "Content", "Publish_date"])  # 添加表头
            wb.save(file_path)
            logging.info(f"Created new file and saved to {file_path}")

        # 追加数据
        wb = load_workbook(file_path)
        ws = wb.active
        for row in data:
            ws.append(row)  # 追加每一行数据
        wb.save(file_path)
        logging.info(f"Appended data to {file_path}")


def enable_proxy():
    os.environ['http_proxy'] = 'http://localhost:10809'
    os.environ['https_proxy'] = 'http://localhost:10809'
    print("全局代理已开启")

def schedule_publish_task():
    enable_proxy()
    item: str
    url_cookies = [(item.split(": ")[0].strip(), item.split(": ")[1].strip()) for item in FileUtils.read_file(os.path.join(os.getcwd(), "file", "cookies.txt"), is_strip=True) if not item.startswith("#")]
    for url, cookies in url_cookies:
        Vip91ChuangYeUtils.publish_vip_91_chuangye_article(url, cookies)



if __name__ == "__main__":

    # 调用静态方法执行操作
    schedule_publish_task()

    # 定义每天定时执行的任务
    schedule.every(1).day.at("08:00").do(schedule_publish_task)  # 设置为每天08:00执行
    while True:
        detect_interval = 60 * 60
        schedule.run_pending()  # 检查是否有任务需要执行
        logging.info(f"等待 {detect_interval} 秒，检测数据。")
        time.sleep(detect_interval)  # 暂停60秒