import os, shutil
from typing import List, Dict, Union, Tuple
import pandas as pd
from openpyxl import Workbook, load_workbook
from .LogUtils import LogUtils


class FileUtils:

    @classmethod
    def clear_directory(cls, directory_path: str) -> None:
        """
        清空指定文件夹

        :param directory_path: 指定文件夹
        :return: None
        """
        # 检查文件夹是否存在
        if os.path.exists(directory_path):
            # 删除整个目录及其子目录
            shutil.rmtree(directory_path)
            # 重新创建空目录
            os.makedirs(directory_path)
            LogUtils.info(f"目录 {directory_path} 已清空")
        else:
            LogUtils.info(f"目录 {directory_path} 不存在")

    @staticmethod
    def copy_files_to_all_subdirs(files_to_copy: List[str], target_dir: str) -> None:
        """
        复制指定文件到指定文件夹及其子孙文件夹下
        :param files_to_copy: 需要复制的文件路径数组
        :param target_dir: 目标文件夹
        :return:
        """
        # 遍历目标文件夹及其所有子文件夹
        for root, dirs, files in os.walk(target_dir):
            # 对每个子目录进行操作
            for file_path in files_to_copy:
                # 获取文件名
                file_name = os.path.basename(file_path)
                # 目标路径
                target_path = os.path.join(root, file_name)

                try:
                    # 复制文件
                    shutil.copy(file_path, target_path)
                    logger.info(f"复制文件 {file_name} 到 {target_path}")
                except Exception as e:
                    LogUtils.error(f"复制文件失败 {file_name}: {e}")

    @staticmethod
    def list_files_in_directory(directory_path) -> list:
        """
        列出指定目录下的所有子文件。
        :param directory_path: 目录的路径。
        :return: 包含目录下所有文件名的列表，如果发生错误则返回空列表。
        """
        try:
            # 列出目标文件夹中的所有项目
            items = os.listdir(directory_path)

            # 过滤出文件
            files = [item for item in items if os.path.isfile(os.path.join(directory_path, item))]
            files.sort()
            return files
        except Exception as e:
            LogUtils.error(f"An error occurred: {e}")
            return []

    @staticmethod
    def list_dirs_in_directory(directory_path: str) -> list:
        """
        列出指定目录下的所有子目录。
        :param directory_path: 目录的路径。
        :return: 包含子目录名称的列表。
        """
        try:
            # 列出目标文件夹中的所有项目
            items = os.listdir(directory_path)

            # 过滤出文件
            dirs = [item for item in items if os.path.isdir(os.path.join(directory_path, item))]

            return dirs
        except Exception as e:
            LogUtils.error(f"An error occurred: {e}")
            return []

    @staticmethod
    def delete_files_with_keyword(directory: str, keyword: str) -> None:
        """
        递归删除指定目录及其子目录中包含特定关键字的文件.

        :param directory: 目标文件夹路径
        :param keyword: 要搜索的文件名关键字
        """
        for root, dirs, files in os.walk(directory):
            for file in files:
                if keyword in file:  # 检查文件名是否包含关键字
                    file_path = os.path.join(root, file)
                    try:
                        os.remove(file_path)
                        LogUtils.info(f"Deleted file: {file_path}")
                    except OSError as e:
                        LogUtils.error(f"Error deleting file {file_path}: {e}")

    @staticmethod
    def rename_images_in_subfolders(folder_path: str) -> None:
        """
        递归地重命名给定文件夹下的所有子文件夹中的图片文件。
        图片文件将被重命名为：子文件夹名_序号.扩展名，其中序号是基于每个子文件夹中的图片数量确定的。

        :param folder_path: 待处理的主文件夹路径。
        """
        for subdir in os.listdir(folder_path):
            subdir_path = os.path.join(folder_path, subdir)

            # 确保当前路径是一个子文件夹
            if os.path.isdir(subdir_path):
                # 获取子文件夹内的所有文件
                images = [f for f in os.listdir(subdir_path) if
                          f.lower().endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp'))]

                # 确定编号的位数，位数由图片的数量决定
                num_images = len(images)
                num_digits = len(str(num_images))  # 序号需要的位数

                # 遍历每个图片文件，重命名
                for idx, image in enumerate(images, start=1):
                    old_image_path = os.path.join(subdir_path, image)

                    # 构造新文件名：子文件夹名_序号.扩展名
                    new_image_name = f"{subdir}_{str(idx).zfill(num_digits)}{os.path.splitext(image)[1]}"
                    new_image_path = os.path.join(subdir_path, new_image_name)

                    # 重命名文件
                    os.rename(old_image_path, new_image_path)
                    LogUtils.info(f"Renamed: {old_image_path} -> {new_image_path}")

    @staticmethod
    def read_file(path: str, is_strip: bool = False, mode: str = 'r', encoding: str = 'u8') -> Union[List[str], None]:
        """
        读取文件，返回文件内容列表。
        :param path: 文件路径。
        :param is_strip: 是否去除每行首尾空白符。
        :param mode: 文件打开模式。
        :param encoding: 文件编码。
        :return: 文件内容列表。
        """
        if path is not None and len(path.strip()) != 0 and os.path.exists(path):
            try:
                with open(path, mode=mode, encoding=encoding) as rf:
                    data_list = rf.readlines()
                if is_strip:
                    data_list = [data.strip() for data in data_list]
                return data_list
            except BaseException as e:
                print(f"FileUtils.read_file(): {e}")
                return None
        else:
            return None

    @staticmethod
    def write_file(path: str, data_list: List[str], mode: str = 'w', encoding: str = 'u8') -> None:
        """
        将数据列表写入文件。
        :param path: 文件路径。
        :param data_list: 要写入的数据列表。
        :param mode: 文件打开模式。
        :param encoding: 文件编码。
        """
        if path is not None and len(path.strip()) != 0:
            parent_path = os.path.split(path)[0]
            if not os.path.exists(parent_path):
                os.makedirs(parent_path)
            try:
                write_list = []
                for data in data_list:
                    if not data.endswith("\n"):
                        data = data + '\n'
                    write_list.append(data)

                with open(path, mode=mode, encoding=encoding) as wf:
                    wf.writelines(write_list)
            except BaseException as e:
                print(f"FileUtils.read_file(): {e}")

    @staticmethod
    def read_xlsx(file_path: str, field_mapping: dict, has_header: bool = True) -> List[Dict]:
        """
        读取本地xlsx文件，并根据field_mapping进行字段映射

        :param file_path: str, xlsx文件路径
        :param field_mapping: dict, 文件表头与字段映射关系
        :param has_header: bool, 文件是否有表头
        :return: List[Dict], 映射后的数据列表
        """
        # 读取Excel文件
        if has_header:
            df = pd.read_excel(file_path, header=0)
        else:
            df = pd.read_excel(file_path, header=None)
            # 手动设置列名为0, 1, 2, ...
            df.columns = range(df.shape[1])

        # 映射字段
        df.rename(columns=field_mapping, inplace=True)

        # 将DataFrame转换为List[Dict]
        data_list = df.to_dict(orient='records')

        return data_list

    @classmethod
    def append_to_excel(cls, file_path: str, data: Union[list, tuple], headers: tuple = None) -> None:
        """
        将数据写入指定 xlsx 文件中。
        规范数据传入，如果插入多行则 data: List[tuple]，单行则 data: tupe

        :param file_path: 文件路径
        :param data: 数据列表
        :param headers: 自定义表头
        :return: None
        """
        if not os.path.exists(file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "TgArticle"
            if headers is not None:
                ws.append(headers)
            wb.save(file_path)

        wb = load_workbook(file_path)
        ws = wb.active
        if isinstance(data, list) or isinstance(data, List):
            for item in data:
                ws.append(item)
        elif isinstance(data, tuple) or isinstance(data, Tuple):
            ws.append(data)
        wb.save(file_path)


if __name__ == '__main__':
    # test_list_files_in_directory()
    vip_links = []
    for i in range(1, 2423):
        vip_links.append(f"https://t.me/zh_vip/{i}")
    FileUtils.write_file(r"C:\Users\Administrator\Desktop\vip_links.txt", vip_links)
